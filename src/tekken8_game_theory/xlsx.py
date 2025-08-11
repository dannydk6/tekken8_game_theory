"""Create input and output excel files for GTO"""

import os
import json
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .convert import json_to_gbt, excel_to_json
from .gto import create_gto_distribution
from .utils import number_to_letter, full_path

def excel_nformat_precision(precision):
    """
    Create an excel number format string based on precision.
    """
    if precision <= 0:
        return '0%'
    s = '0.'
    for i in range(1, precision+1):
        s+='0'
    return s+'%'

def create_excel_file(out_filename, input_dir, output_dir=None, precision=0, drop_zeros=True):
    """
    Create an excel file which contains all the nash equilibrium payoffs computed for .gbt files in input_dir.
    
    Args:
        out_filename (str): Name of the output file.
        input_dir (str): Folder containing .gbt files
        output_dir (str|Optional): optional output directory for the excel file
        precision (int): Precision for float values.
        drop_zeros (bool): Whether to drop zeros from distributions.
    
    Returns:
        Writes excel file.
    """
    files = os.listdir(input_dir)
    
    gto_distributions = []
    
    # Create GTO Distributions
    for filename in files:
        if filename.endswith(".gbt"):
            df, metadata = create_gto_distribution(filename, 
                                                   base_path=input_dir, 
                                                   precision=precision, 
                                                   drop_zeros=drop_zeros,
                                                   no_prec_percent=True)
            gto_distributions.append({'df': df, 'metadata': metadata})
    
    out_path = full_path(out_filename, output_dir)
    
    # Define sleek styles
    title_font = Font(size=18, bold=True, color="000000")  # Black, bold, large font
    metadata_key_font = Font(size=12, bold=True, color="333333")  # Dark gray for keys
    metadata_value_font = Font(size=12, color="000000")  # Black for values
    header_font = Font(size=12, bold=True, color="FFFFFF")  # White for table headers
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Light blue
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    title_border_style = Border(
        bottom=Side(style="thin")
    )
    alignment_center = Alignment(horizontal="center", vertical="center")
    
    # Define Yellow Neutral Style
    yellow_neutral_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Pale yellow
    yellow_neutral_font = Font(size=12, color="7F6000")  # Dark brownish-gray

    # Define Green Good Style
    green_good_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE" if False else "C6EFCE", end_color="C6EFCE")
    green_good_font = Font(size=12, color="006100")

    # Define Red Bad style
    red_bad_fill = PatternFill(fill_type="solid", start_color="FFFFC7CE", end_color="FFFFC7CE")
    red_bad_font = Font(size=12, color="FF9C0006")
    
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for gto_distribution in gto_distributions:
            df = gto_distribution["df"]
            metadata = gto_distribution["metadata"]

            # Get the values for P1 and P2
            P1 = metadata['P1']
            P2 = metadata['P2']

            #print(f"P1: {P1}\nP2: {P2}\n")
            
            # Sheet name
            sheet_name = metadata['title'][:min(31,len(metadata['title']))]
            
            #print(df)
            #print(metadata)
            
            # Starting row for gto data
            startrow = len(metadata.keys())+2

            df.to_excel(writer, index=False, startrow=startrow, sheet_name=sheet_name)

            workbook = writer.book
            sheet = writer.sheets[sheet_name]

            # Last column of dataset
            ending_col = number_to_letter(len(df.columns))

            # Write the title
            sheet.merge_cells(f"A1:{ending_col}1")  # Merge cells for the title
            sheet["A1"] = metadata["title"]
            sheet["A1"].font = title_font
            sheet["A1"].alignment = alignment_center

            # Add border to title
            for l in range(1, len(df.columns)+1):
                sheet[f"{number_to_letter(l)}1"].border = title_border_style

            row = 2
            metadata.pop('title', None)
            for key, value in metadata.items():
                sheet[f"A{row}"] = key
                sheet[f"A{row}"].font = metadata_key_font
                sheet[f"A{row}"].alignment = Alignment(wrap_text=True)
                sheet[f"B{row}"] = value
                sheet[f"B{row}"].font = metadata_value_font
                sheet[f"B{row}"].alignment = Alignment(wrap_text=True)
                row += 1

            # Apply sleek styles to the table header
            for cell in sheet[startrow+1]:  # Header row is the 6th row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = border_style

            # Apply borders to the data table and Adjust precision on distribution column
            fmt_string = excel_nformat_precision(precision)
            for row in sheet.iter_rows(min_row=startrow+1, max_row=startrow + len(df) + 1, min_col=1, max_col=len(df.columns)):
                # Get the current player's name
                player_val = row[0].value

                for cell in row:
                    cell.border = border_style
                    cell.alignment = Alignment(wrap_text=True)

                    #print(f"player_val: {player_val}, P1: {P1}, P2: {P2}")

                    # Change the fill if the row contains P1 or P2
                    if player_val == P1:
                        cell.font = green_good_font
                        cell.fill = green_good_fill
                    elif player_val == P2:
                        cell.font = red_bad_font
                        cell.fill = red_bad_fill

                    # Distribution column formatting
                    if cell.column == 3:
                        cell.number_format = fmt_string

            # Adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)  # Get the column letter
                for cell in col:
                    try:  # Necessary to handle cases where the cell value is None
                        # Excluding the "Comment" field
                        # TODO: replace hardcoded B5 with ref to comment cell
                        if cell.value and cell.coordinate != 'B5':
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                sheet.column_dimensions[col_letter].width = max_length + 2  # Add some padding

            # Highlight P1 Cell
            p1_cell = "B2"
            sheet[p1_cell].fill = green_good_fill
            sheet[p1_cell].font = green_good_font
            sheet[p1_cell].border = border_style

            # Highlight P2 Cell
            p2_cell = "B3"
            sheet[p2_cell].fill = red_bad_fill
            sheet[p2_cell].font = red_bad_font
            sheet[p2_cell].border = border_style

            # Highlight Payoff
            payoff_cell = "B4"
            sheet[payoff_cell].fill = yellow_neutral_fill
            sheet[payoff_cell].font = yellow_neutral_font
            sheet[payoff_cell].border = border_style

    return gto_distributions

def full_run(out_filename, 
             input_dir,
             artifacts_dir,
             output_dir=None, 
             precision=0, 
             drop_zeros=True):
    """
    Create an excel file which contains all the nash equilibrium payoffs computed for .gbt files 
    in artifacts_dir. The .gbt files are obtained by converting json and .xlsx files from input_dir
    
    Args:
        out_filename (str): Name of the output file.
        input_dir (str): Folder containing .xlsx and/or .json files.
        artifacts_dir (str): Folder containing .gbt files.
        output_dir (str|Optional): optional output directory for the excel file
        precision (int): Precision for float values.
        drop_zeros (bool): Whether to drop zeros from distributions.
    
    Returns:
        Writes excel file.
    """
    # Files in input_dir
    scenario_files = os.listdir(input_dir)
    
    # Iterate over input files  
    for scenario_file in scenario_files:

        file_extension = ''
        if scenario_file.endswith('.json'):
            file_extension = '.json'
        elif scenario_file.endswith('.xlsx'):
            file_extension = '.xlsx'
        else:
            print(f'Warning: {scenario_file} is not a .json or .xlsx and will be ignored.')
            continue

        # Set path for .gbt converted file based on artifacts_dir
        out_file = scenario_file.replace(file_extension, '.gbt')
        out_path = f'{artifacts_dir}/{out_file}'

        js=''
        if file_extension == '.json':
            with open(f'{input_dir}/{scenario_file}', 'r') as f:
                js = json.load(f)
        else:
            js = excel_to_json(scenario_file, base_path=input_dir)

        gbt = json_to_gbt(js, out_path)
    
    # Create final excel file
    gto_distributions = create_excel_file(out_filename, 
                                          input_dir=artifacts_dir, 
                                          output_dir=output_dir, 
                                          precision=precision, 
                                          drop_zeros=drop_zeros)
    
    return gto_distributions
